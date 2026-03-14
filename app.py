import streamlit as st
import pandas as pd
import numpy as np
import re
import io
from datetime import datetime

# ─── PAGE CONFIG ────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="DataClean Pro · ETL Inteligente",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─── CUSTOM CSS ─────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Sora:wght@300;400;500;600;700&family=JetBrains+Mono:wght@400;500&display=swap');

html, body, [class*="css"] {
    font-family: 'Sora', sans-serif;
}

/* Background */
.stApp {
    background: #0a0e1a;
    color: #e2e8f0;
}

/* Sidebar */
[data-testid="stSidebar"] {
    background: #0f1528 !important;
    border-right: 1px solid #1e2a45;
}
[data-testid="stSidebar"] .stMarkdown h1,
[data-testid="stSidebar"] .stMarkdown h2,
[data-testid="stSidebar"] .stMarkdown h3 {
    color: #7dd3fc;
}

/* Main header */
.main-header {
    background: linear-gradient(135deg, #1e2a45 0%, #0f1528 100%);
    border: 1px solid #2563eb33;
    border-radius: 16px;
    padding: 32px 40px;
    margin-bottom: 28px;
    position: relative;
    overflow: hidden;
}
.main-header::before {
    content: '';
    position: absolute;
    top: -50%;
    right: -10%;
    width: 400px;
    height: 400px;
    background: radial-gradient(circle, #2563eb22 0%, transparent 70%);
    border-radius: 50%;
}
.main-header h1 {
    font-size: 2.4rem;
    font-weight: 700;
    color: #f8fafc;
    margin: 0 0 8px 0;
    letter-spacing: -0.02em;
}
.main-header .tagline {
    color: #7dd3fc;
    font-size: 1rem;
    font-weight: 400;
    margin: 0;
}
.badge {
    display: inline-block;
    background: #2563eb22;
    border: 1px solid #2563eb55;
    color: #7dd3fc;
    padding: 4px 12px;
    border-radius: 20px;
    font-size: 0.75rem;
    font-weight: 500;
    letter-spacing: 0.05em;
    text-transform: uppercase;
    margin-bottom: 16px;
}

/* Cards */
.metric-card {
    background: #131929;
    border: 1px solid #1e2a45;
    border-radius: 12px;
    padding: 20px 24px;
    text-align: center;
    transition: border-color 0.2s;
}
.metric-card:hover { border-color: #2563eb55; }
.metric-card .value {
    font-size: 2rem;
    font-weight: 700;
    color: #7dd3fc;
    font-family: 'JetBrains Mono', monospace;
}
.metric-card .label {
    font-size: 0.8rem;
    color: #64748b;
    text-transform: uppercase;
    letter-spacing: 0.08em;
    margin-top: 4px;
}

/* Steps */
.step-container {
    background: #131929;
    border: 1px solid #1e2a45;
    border-left: 3px solid #2563eb;
    border-radius: 8px;
    padding: 16px 20px;
    margin: 8px 0;
    font-size: 0.9rem;
}
.step-container.done {
    border-left-color: #10b981;
    background: #0d1f1a;
}
.step-container.done .step-icon { color: #10b981; }
.step-title { font-weight: 600; color: #e2e8f0; margin-bottom: 4px; }
.step-detail { color: #64748b; font-size: 0.82rem; font-family: 'JetBrains Mono', monospace; }

/* Upload zone */
[data-testid="stFileUploader"] {
    background: #131929;
    border: 2px dashed #1e2a45;
    border-radius: 12px;
    transition: border-color 0.2s;
}
[data-testid="stFileUploader"]:hover {
    border-color: #2563eb55;
}

/* Buttons */
.stButton > button {
    background: linear-gradient(135deg, #2563eb, #1d4ed8);
    color: white;
    border: none;
    border-radius: 8px;
    font-family: 'Sora', sans-serif;
    font-weight: 600;
    font-size: 0.95rem;
    padding: 12px 28px;
    width: 100%;
    transition: all 0.2s;
    letter-spacing: 0.01em;
}
.stButton > button:hover {
    background: linear-gradient(135deg, #1d4ed8, #1e40af);
    transform: translateY(-1px);
    box-shadow: 0 4px 20px #2563eb44;
}

/* Download button */
.stDownloadButton > button {
    background: linear-gradient(135deg, #10b981, #059669) !important;
    color: white !important;
    border: none !important;
    border-radius: 8px !important;
    font-family: 'Sora', sans-serif !important;
    font-weight: 600 !important;
    width: 100%;
    padding: 12px 28px !important;
    transition: all 0.2s !important;
}
.stDownloadButton > button:hover {
    transform: translateY(-1px) !important;
    box-shadow: 0 4px 20px #10b98144 !important;
}

/* Tables */
.stDataFrame {
    border: 1px solid #1e2a45;
    border-radius: 8px;
}

/* Alerts */
.stSuccess {
    background: #0d1f1a;
    border: 1px solid #10b981;
    border-radius: 8px;
}
.stError {
    background: #1f0d0d;
    border: 1px solid #ef4444;
    border-radius: 8px;
}
.stInfo {
    background: #0f1a2e;
    border: 1px solid #2563eb;
    border-radius: 8px;
}
.stWarning {
    background: #1f1a0d;
    border: 1px solid #f59e0b;
    border-radius: 8px;
}

/* Selectbox, multiselect */
[data-testid="stSelectbox"], [data-testid="stMultiSelect"] {
    background: #131929;
}

/* Section headers */
.section-title {
    font-size: 1.1rem;
    font-weight: 600;
    color: #e2e8f0;
    margin: 24px 0 12px 0;
    padding-bottom: 8px;
    border-bottom: 1px solid #1e2a45;
}

/* Log box */
.log-box {
    background: #080c14;
    border: 1px solid #1e2a45;
    border-radius: 8px;
    padding: 16px;
    font-family: 'JetBrains Mono', monospace;
    font-size: 0.8rem;
    color: #7dd3fc;
    max-height: 300px;
    overflow-y: auto;
    line-height: 1.8;
}

/* Hide streamlit branding */
#MainMenu {visibility: hidden;}
footer {visibility: hidden;}
header {visibility: hidden;}
</style>
""", unsafe_allow_html=True)


# ─── ETL ENGINE ─────────────────────────────────────────────────────────────

def extrair_cidade(endereco: str) -> str:
    """Extrai cidade da string de endereço: Rua, Número, Cidade, Estado, País"""
    if pd.isna(endereco) or str(endereco).strip() == "":
        return ""
    partes = [p.strip() for p in str(endereco).split(",")]
    return partes[2] if len(partes) >= 4 else ""


def normalizar_nome(nome: str) -> str:
    """Converte 'SOBRENOME, Nome' → 'Nome Sobrenome' em Proper Case"""
    if pd.isna(nome) or str(nome).strip() == "":
        return ""
    nome = str(nome)
    if "," in nome:
        partes = nome.split(",", 1)
        sobrenome = partes[0].strip()
        primeiro = partes[1].strip()
        return f"{primeiro} {sobrenome}".title()
    return nome.title()


def mapear_genero(g: str) -> str:
    mapa = {"M": "Masculino", "F": "Feminino"}
    return mapa.get(str(g).strip().upper(), g)


def detectar_coluna(df: pd.DataFrame, candidatos: list) -> str | None:
    """Detecta automaticamente o nome de uma coluna por lista de candidatos."""
    cols_lower = {c.lower(): c for c in df.columns}
    for cand in candidatos:
        if cand.lower() in cols_lower:
            return cols_lower[cand.lower()]
    return None


def rodar_etl(
    arquivo_bytes: bytes,
    nome_arquivo: str,
    skip_rows: int,
    config: dict,
    log: list,
) -> tuple[pd.DataFrame | None, list]:

    log.append(f"📂 Arquivo recebido: <b>{nome_arquivo}</b>")

    # ── 1. LEITURA ────────────────────────────────────────────────────────
    try:
        ext = nome_arquivo.split(".")[-1].lower()
        if ext in ("xlsx", "xls"):
            df = pd.read_excel(
                io.BytesIO(arquivo_bytes),
                skiprows=skip_rows,
                header=0,
                engine="openpyxl" if ext == "xlsx" else None,
            )
        elif ext == "csv":
            df = pd.read_csv(io.BytesIO(arquivo_bytes), skiprows=skip_rows, header=0)
        else:
            log.append("❌ Formato não suportado.")
            return None, log
        log.append(f"✅ Leitura OK — {df.shape[0]} linhas × {df.shape[1]} colunas")
    except Exception as e:
        log.append(f"❌ Erro na leitura: {e}")
        return None, log

    # ── 2. LIMPEZA ESTRUTURAL ─────────────────────────────────────────────
    antes = df.shape[0]
    df.dropna(how="all", inplace=True)                     # linhas 100% vazias
    df.dropna(axis=1, how="all", inplace=True)             # colunas 100% vazias
    df.columns = df.columns.astype(str).str.strip()        # limpar nomes
    log.append(f"🧹 Linhas vazias removidas: {antes - df.shape[0]}")

    # Remover coluna de Observações
    col_obs = detectar_coluna(df, config.get("col_observacoes", ["Observações", "Observacoes", "Obs", "Notas", "Notes"]))
    if col_obs:
        df.drop(columns=[col_obs], inplace=True)
        log.append(f"🗑️  Coluna removida: {col_obs}")

    # ── 3. SPLIT PRODUTO → Produto + Marca ───────────────────────────────
    col_prod = detectar_coluna(df, config.get("col_produto", ["Produto", "Product", "Item"]))
    if col_prod:
        delimitador = config.get("delimitador_produto", " - ")
        split = df[col_prod].astype(str).str.split(delimitador, n=1, expand=True)
        df["Produto"] = split[0].str.strip()
        df["Marca"] = split[1].str.strip() if split.shape[1] > 1 else ""
        if col_prod not in ("Produto",):
            df.drop(columns=[col_prod], inplace=True)
        log.append(f"✂️  Split de produto: '{col_prod}' → Produto + Marca")
    else:
        log.append("⚠️  Coluna de produto não encontrada — pulando split")

    # ── 4. NORMALIZAÇÃO DE NOMES ─────────────────────────────────────────
    col_cli = detectar_coluna(df, config.get("col_cliente", ["Cliente", "Client", "Nome", "Name"]))
    if col_cli:
        df[col_cli] = df[col_cli].apply(normalizar_nome)
        log.append(f"👤 Nomes normalizados: {col_cli}")
    else:
        log.append("⚠️  Coluna de cliente não encontrada — pulando normalização")

    # ── 5. MAPEAMENTO DE GÊNERO ──────────────────────────────────────────
    col_gen = detectar_coluna(df, config.get("col_genero", ["Gênero", "Genero", "Gender", "Sexo"]))
    if col_gen:
        df[col_gen] = df[col_gen].apply(mapear_genero)
        log.append(f"🏷️  Gênero mapeado: {col_gen}")
    else:
        log.append("⚠️  Coluna de gênero não encontrada — pulando mapeamento")

    # ── 6. EXTRAÇÃO DE CIDADE ─────────────────────────────────────────────
    col_end = detectar_coluna(df, config.get("col_endereco", ["Endereço", "Endereco", "Address", "Logradouro"]))
    if col_end:
        df["Cidade"] = df[col_end].apply(extrair_cidade)
        log.append(f"🏙️  Cidade extraída de: {col_end}")
    else:
        log.append("⚠️  Coluna de endereço não encontrada — pulando extração de cidade")

    # ── 7. TIPAGEM ────────────────────────────────────────────────────────
    colunas_data = config.get("colunas_data", [])
    colunas_valor = config.get("colunas_valor", [])

    for col in colunas_data:
        col_real = detectar_coluna(df, [col])
        if col_real:
            try:
                df[col_real] = pd.to_datetime(df[col_real], dayfirst=True, errors="coerce")
                log.append(f"📅 Tipagem datetime: {col_real}")
            except Exception:
                pass

    for col in colunas_valor:
        col_real = detectar_coluna(df, [col])
        if col_real:
            try:
                df[col_real] = (
                    df[col_real].astype(str)
                    .str.replace(r"[R$\s\.]", "", regex=True)
                    .str.replace(",", ".", regex=False)
                )
                df[col_real] = pd.to_numeric(df[col_real], errors="coerce").astype("float64")
                log.append(f"💰 Tipagem float: {col_real}")
            except Exception:
                pass

    # ── 8. RESET INDEX ────────────────────────────────────────────────────
    df.reset_index(drop=True, inplace=True)
    log.append(f"🎉 ETL concluído! Base final: <b>{df.shape[0]} linhas × {df.shape[1]} colunas</b>")

    return df, log


def gerar_excel(df: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Base_Limpa")
        ws = writer.sheets["Base_Limpa"]
        # Formatar cabeçalhos
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        header_fill = PatternFill("solid", start_color="1E3A5F")
        header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
        thin = Side(style="thin", color="2A4A7F")
        border = Border(bottom=Side(style="medium", color="2563EB"))
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        # Auto-width
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)
        ws.row_dimensions[1].height = 22
    return buf.getvalue()


# ─── SIDEBAR ────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("""
    <div style="padding:16px 0 8px 0">
        <div style="font-size:1.4rem;font-weight:700;color:#f8fafc;letter-spacing:-0.02em">⚡ DataClean Pro</div>
        <div style="font-size:0.75rem;color:#475569;margin-top:4px">ETL Inteligente para Contabilidade</div>
    </div>
    """, unsafe_allow_html=True)

    st.divider()
    st.markdown("### ⚙️ Configuração do ETL")

    skip_rows = st.number_input(
        "Linhas de cabeçalho para pular",
        min_value=0, max_value=20, value=3,
        help="Quantas linhas inúteis existem antes dos dados reais?"
    )

    st.markdown("**Nomes das colunas-chave** (separados por vírgula)")
    col_produto_input = st.text_input("Coluna Produto", value="Produto")
    col_cliente_input = st.text_input("Coluna Cliente", value="Cliente")
    col_genero_input  = st.text_input("Coluna Gênero",  value="Gênero")
    col_endereco_input = st.text_input("Coluna Endereço", value="Endereço")
    col_obs_input = st.text_input("Coluna Observações (remover)", value="Observações")
    delimitador = st.text_input("Delimitador Produto→Marca", value=" - ")

    st.markdown("**Colunas de Data** (separadas por vírgula)")
    datas_input = st.text_input("", value="Data, Data Venda, Data Pedido, Data Nascimento")

    st.markdown("**Colunas Numéricas/Valor** (separadas por vírgula)")
    valores_input = st.text_input("", value="Valor, Preço, Total, Quantidade")

    st.divider()
    st.markdown("""
    <div style="font-size:0.75rem;color:#475569;line-height:1.8">
    📋 <b>Transformações aplicadas:</b><br>
    ✔ Pular linhas de cabeçalho<br>
    ✔ Remover linhas/colunas vazias<br>
    ✔ Split Produto → Produto + Marca<br>
    ✔ Normalizar SOBRENOME, Nome<br>
    ✔ M/F → Masculino/Feminino<br>
    ✔ Extrair Cidade do Endereço<br>
    ✔ Tipagem datetime e float
    </div>
    """, unsafe_allow_html=True)


# ─── MAIN CONTENT ───────────────────────────────────────────────────────────

st.markdown("""
<div class="main-header">
    <div class="badge">⚡ Automação ETL · v1.0</div>
    <h1>DataClean Pro</h1>
    <p class="tagline">Transforme planilhas brutas em bases prontas para análise — em segundos.</p>
</div>
""", unsafe_allow_html=True)

# Montagem da config a partir do sidebar
config = {
    "col_produto":    [c.strip() for c in col_produto_input.split(",") if c.strip()],
    "col_cliente":    [c.strip() for c in col_cliente_input.split(",") if c.strip()],
    "col_genero":     [c.strip() for c in col_genero_input.split(",") if c.strip()],
    "col_endereco":   [c.strip() for c in col_endereco_input.split(",") if c.strip()],
    "col_observacoes":[c.strip() for c in col_obs_input.split(",") if c.strip()],
    "delimitador_produto": delimitador,
    "colunas_data":   [c.strip() for c in datas_input.split(",") if c.strip()],
    "colunas_valor":  [c.strip() for c in valores_input.split(",") if c.strip()],
}

# ── UPLOAD ──────────────────────────────────────────────────────────────────
col_up, col_info = st.columns([2, 1])

with col_up:
    st.markdown('<div class="section-title">📁 Carregar Arquivo</div>', unsafe_allow_html=True)
    uploaded_file = st.file_uploader(
        "Arraste seu arquivo Excel ou CSV aqui",
        type=["xlsx", "xls", "csv"],
        label_visibility="collapsed",
    )

with col_info:
    st.markdown('<div class="section-title">📋 Pipeline ETL</div>', unsafe_allow_html=True)
    etapas = [
        ("1", "Limpeza Estrutural", "Remove linhas/colunas vazias e cabeçalho sujo"),
        ("2", "Split de Produto", f"'{col_produto_input}' → Produto + Marca"),
        ("3", "Normalização", "'SOBRENOME, Nome' → 'Nome Sobrenome'"),
        ("4", "Mapeamento", "M → Masculino · F → Feminino"),
        ("5", "Extração de Cidade", "Cidade extraída do campo Endereço"),
        ("6", "Tipagem", "Datas e valores com tipos corretos"),
    ]
    for num, titulo, detalhe in etapas:
        st.markdown(f"""
        <div class="step-container">
            <div class="step-title">Etapa {num} · {titulo}</div>
            <div class="step-detail">{detalhe}</div>
        </div>
        """, unsafe_allow_html=True)

# ── PROCESSAMENTO ────────────────────────────────────────────────────────────
if uploaded_file:
    arquivo_bytes = uploaded_file.read()
    nome_arquivo = uploaded_file.name

    st.divider()
    col_btn, col_empty = st.columns([1, 2])
    with col_btn:
        processar = st.button("▶ Processar Arquivo", use_container_width=True)

    if processar or st.session_state.get("processado"):
        if processar:
            log = []
            with st.spinner("Processando ETL..."):
                df_limpo, log = rodar_etl(arquivo_bytes, nome_arquivo, skip_rows, config, log)
            st.session_state["df_limpo"] = df_limpo
            st.session_state["log"] = log
            st.session_state["processado"] = True
            st.session_state["nome_arquivo"] = nome_arquivo

        df_limpo = st.session_state.get("df_limpo")
        log      = st.session_state.get("log", [])

        # Log
        st.markdown('<div class="section-title">📋 Log de Processamento</div>', unsafe_allow_html=True)
        log_html = "<br>".join(f"› {linha}" for linha in log)
        st.markdown(f'<div class="log-box">{log_html}</div>', unsafe_allow_html=True)

        if df_limpo is not None and not df_limpo.empty:
            st.divider()

            # Métricas
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.markdown(f'<div class="metric-card"><div class="value">{df_limpo.shape[0]:,}</div><div class="label">Linhas</div></div>', unsafe_allow_html=True)
            with col2:
                st.markdown(f'<div class="metric-card"><div class="value">{df_limpo.shape[1]}</div><div class="label">Colunas</div></div>', unsafe_allow_html=True)
            with col3:
                nulos = int(df_limpo.isnull().sum().sum())
                st.markdown(f'<div class="metric-card"><div class="value">{nulos}</div><div class="label">Valores Nulos</div></div>', unsafe_allow_html=True)
            with col4:
                dupes = int(df_limpo.duplicated().sum())
                st.markdown(f'<div class="metric-card"><div class="value">{dupes}</div><div class="label">Duplicatas</div></div>', unsafe_allow_html=True)

            st.divider()

            # Preview
            st.markdown('<div class="section-title">👁️ Preview da Base Limpa</div>', unsafe_allow_html=True)

            col_prev, col_dl = st.columns([3, 1])
            with col_prev:
                n_linhas = st.slider("Linhas para visualizar", 5, min(100, len(df_limpo)), 10)
            with col_dl:
                st.markdown("<br>", unsafe_allow_html=True)
                excel_bytes = gerar_excel(df_limpo)
                st.download_button(
                    label="⬇ Baixar Base_Vendas_Limpa.xlsx",
                    data=excel_bytes,
                    file_name="Base_Vendas_Limpa.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

            st.dataframe(
                df_limpo.head(n_linhas),
                use_container_width=True,
                height=400,
            )

            # Tipos das colunas
            with st.expander("🔍 Tipos de dados das colunas"):
                tipos = df_limpo.dtypes.reset_index()
                tipos.columns = ["Coluna", "Tipo"]
                tipos["Tipo"] = tipos["Tipo"].astype(str)
                st.dataframe(tipos, use_container_width=True, hide_index=True)

            # Estatísticas
            with st.expander("📊 Estatísticas das colunas numéricas"):
                nums = df_limpo.select_dtypes(include="number")
                if not nums.empty:
                    st.dataframe(nums.describe().round(2), use_container_width=True)
                else:
                    st.info("Nenhuma coluna numérica detectada.")

        else:
            st.error("Não foi possível processar o arquivo. Verifique as configurações no painel lateral.")

else:
    # Estado vazio
    st.session_state.pop("processado", None)
    st.session_state.pop("df_limpo", None)

    st.markdown("<br>", unsafe_allow_html=True)
    st.info("👆 Faça o upload de um arquivo Excel (.xlsx / .xls) ou CSV para começar.")

    # Cards de funcionalidades
    st.markdown('<div class="section-title">🚀 O que o DataClean Pro faz por você</div>', unsafe_allow_html=True)
    col1, col2, col3 = st.columns(3)

    features = [
        ("🧹", "Limpeza Automática", "Remove cabeçalhos sujos, linhas e colunas em branco sem nenhuma configuração manual."),
        ("✂️", "Split Inteligente", "Separa campos aglutinados como 'Produto - Marca' em colunas individuais."),
        ("👤", "Normalização de Nomes", "Converte 'SOBRENOME, Nome' para o formato legível 'Nome Sobrenome' automaticamente."),
        ("🏙️", "Extração de Entidade", "Extrai cidades de strings de endereço longas com precisão."),
        ("📅", "Tipagem Correta", "Garante que datas sejam datetime e valores sejam float — essencial para análises."),
        ("⬇️", "Exportação Formatada", "Gera um .xlsx profissional com cabeçalhos formatados e colunas auto-dimensionadas."),
    ]

    cols = [col1, col2, col3, col1, col2, col3]
    for i, (icon, titulo, desc) in enumerate(features):
        with cols[i]:
            st.markdown(f"""
            <div class="metric-card" style="text-align:left;padding:20px">
                <div style="font-size:1.6rem;margin-bottom:8px">{icon}</div>
                <div style="font-weight:600;color:#e2e8f0;margin-bottom:6px">{titulo}</div>
                <div style="font-size:0.82rem;color:#64748b;line-height:1.6">{desc}</div>
            </div>
            """, unsafe_allow_html=True)
            st.markdown("<br>", unsafe_allow_html=True)