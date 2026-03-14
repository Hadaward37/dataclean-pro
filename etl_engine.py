"""
etl_engine.py
─────────────
Motor ETL standalone — pode ser importado pelo app Streamlit,
rodado via CLI, usado como AWS Lambda ou Google Cloud Function,
ou integrado a um Watchdog de pasta.

Uso via CLI:
    python etl_engine.py caminho/para/arquivo.xlsx

Uso como módulo:
    from etl_engine import ETLConfig, rodar_pipeline
    config = ETLConfig()
    df, log = rodar_pipeline(bytes_arquivo, "vendas.xlsx", config)
"""

from __future__ import annotations
import io
import re
import sys
import json
import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import pandas as pd

logger = logging.getLogger(__name__)


# ─── CONFIGURAÇÃO ─────────────────────────────────────────────────────────

@dataclass
class ETLConfig:
    """Configuração completa do pipeline ETL. Todos os campos têm defaults sensatos."""

    # Quantas linhas pular antes do cabeçalho real
    skip_rows: int = 3

    # Candidatos de nome de coluna (detecta automaticamente)
    col_produto:     list[str] = field(default_factory=lambda: ["Produto", "Product", "Item"])
    col_cliente:     list[str] = field(default_factory=lambda: ["Cliente", "Client", "Nome", "Name"])
    col_genero:      list[str] = field(default_factory=lambda: ["Gênero", "Genero", "Gender", "Sexo"])
    col_endereco:    list[str] = field(default_factory=lambda: ["Endereço", "Endereco", "Address"])
    col_observacoes: list[str] = field(default_factory=lambda: ["Observações", "Observacoes", "Obs", "Notas"])

    # Delimitador usado em "Produto - Marca"
    delimitador_produto: str = " - "

    # Colunas a converter para datetime / float
    colunas_data:  list[str] = field(default_factory=lambda: ["Data", "Data Venda", "Data Pedido", "Data Nascimento"])
    colunas_valor: list[str] = field(default_factory=lambda: ["Valor", "Preço", "Total", "Quantidade"])

    # Mapeamento de gênero
    mapa_genero: dict[str, str] = field(default_factory=lambda: {"M": "Masculino", "F": "Feminino"})

    @classmethod
    def from_json(cls, path: str) -> "ETLConfig":
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        return cls(**data)


# ─── HELPERS ──────────────────────────────────────────────────────────────

def _detectar_coluna(df: pd.DataFrame, candidatos: list[str]) -> Optional[str]:
    """Retorna o primeiro nome de coluna que bater com qualquer candidato (case-insensitive)."""
    cols_lower = {c.lower(): c for c in df.columns}
    for cand in candidatos:
        if cand.lower() in cols_lower:
            return cols_lower[cand.lower()]
    return None


def _extrair_cidade(endereco: str) -> str:
    """Rua, Número, Cidade, Estado, País → Cidade (índice 2)."""
    if pd.isna(endereco) or not str(endereco).strip():
        return ""
    partes = [p.strip() for p in str(endereco).split(",")]
    return partes[2] if len(partes) >= 4 else ""


def _normalizar_nome(nome: str) -> str:
    """'SANTOS, Carlos' → 'Carlos Santos'."""
    if pd.isna(nome) or not str(nome).strip():
        return ""
    nome = str(nome)
    if "," in nome:
        sobrenome, primeiro = nome.split(",", 1)
        return f"{primeiro.strip()} {sobrenome.strip()}".title()
    return nome.title()


def _limpar_numero(valor) -> Optional[float]:
    """Remove R$, pontos de milhar e converte vírgula decimal → float."""
    try:
        s = re.sub(r"[R$\s\.]", "", str(valor)).replace(",", ".")
        return float(s)
    except (ValueError, TypeError):
        return None


# ─── PIPELINE PRINCIPAL ────────────────────────────────────────────────────

def rodar_pipeline(
    arquivo_bytes: bytes,
    nome_arquivo: str,
    config: Optional[ETLConfig] = None,
) -> tuple[Optional[pd.DataFrame], list[str]]:
    """
    Executa o pipeline ETL completo.

    Parâmetros
    ----------
    arquivo_bytes : conteúdo do arquivo em bytes
    nome_arquivo  : nome original (para detectar extensão)
    config        : ETLConfig — usa defaults se None

    Retorna
    -------
    (DataFrame limpo | None, lista de mensagens de log)
    """
    cfg = config or ETLConfig()
    log: list[str] = []

    # ── ETAPA 1: LEITURA ────────────────────────────────────────────────
    log.append(f"[INFO] Arquivo recebido: {nome_arquivo}")
    ext = Path(nome_arquivo).suffix.lower()

    try:
        if ext in (".xlsx", ".xls"):
            df = pd.read_excel(
                io.BytesIO(arquivo_bytes),
                skiprows=cfg.skip_rows,
                header=0,
                engine="openpyxl" if ext == ".xlsx" else None,
            )
        elif ext == ".csv":
            df = pd.read_csv(io.BytesIO(arquivo_bytes), skiprows=cfg.skip_rows, header=0)
        else:
            log.append("[ERRO] Formato não suportado. Use .xlsx, .xls ou .csv")
            return None, log

        log.append(f"[OK] Leitura concluída: {df.shape[0]} linhas × {df.shape[1]} colunas")
    except Exception as e:
        log.append(f"[ERRO] Falha na leitura: {e}")
        return None, log

    # ── ETAPA 2: LIMPEZA ESTRUTURAL ─────────────────────────────────────
    n_antes = df.shape[0]
    df.dropna(how="all", inplace=True)
    df.dropna(axis=1, how="all", inplace=True)
    df.columns = df.columns.astype(str).str.strip()
    log.append(f"[OK] Linhas 100%% vazias removidas: {n_antes - df.shape[0]}")

    col_obs = _detectar_coluna(df, cfg.col_observacoes)
    if col_obs:
        df.drop(columns=[col_obs], inplace=True)
        log.append(f"[OK] Coluna de observações removida: '{col_obs}'")

    # ── ETAPA 3: SPLIT PRODUTO → Produto + Marca ─────────────────────────
    col_prod = _detectar_coluna(df, cfg.col_produto)
    if col_prod:
        split = df[col_prod].astype(str).str.split(cfg.delimitador_produto, n=1, expand=True)
        df["Produto"] = split[0].str.strip()
        df["Marca"]   = split[1].str.strip() if split.shape[1] > 1 else ""
        if col_prod != "Produto":
            df.drop(columns=[col_prod], inplace=True)
        log.append(f"[OK] Produto split em 'Produto' e 'Marca' (delim='{cfg.delimitador_produto}')")
    else:
        log.append("[AVISO] Coluna de produto não encontrada — etapa ignorada")

    # ── ETAPA 4: NORMALIZAÇÃO DE NOMES ──────────────────────────────────
    col_cli = _detectar_coluna(df, cfg.col_cliente)
    if col_cli:
        df[col_cli] = df[col_cli].apply(_normalizar_nome)
        log.append(f"[OK] Nomes normalizados: '{col_cli}'")
    else:
        log.append("[AVISO] Coluna de cliente não encontrada — etapa ignorada")

    # ── ETAPA 5: MAPEAMENTO DE GÊNERO ────────────────────────────────────
    col_gen = _detectar_coluna(df, cfg.col_genero)
    if col_gen:
        df[col_gen] = df[col_gen].astype(str).str.strip().str.upper().map(
            lambda g: cfg.mapa_genero.get(g, g)
        )
        log.append(f"[OK] Gênero mapeado: '{col_gen}'")
    else:
        log.append("[AVISO] Coluna de gênero não encontrada — etapa ignorada")

    # ── ETAPA 6: EXTRAÇÃO DE CIDADE ──────────────────────────────────────
    col_end = _detectar_coluna(df, cfg.col_endereco)
    if col_end:
        df["Cidade"] = df[col_end].apply(_extrair_cidade)
        log.append(f"[OK] Coluna 'Cidade' extraída de '{col_end}'")
    else:
        log.append("[AVISO] Coluna de endereço não encontrada — etapa ignorada")

    # ── ETAPA 7: TIPAGEM ─────────────────────────────────────────────────
    for col in cfg.colunas_data:
        col_real = _detectar_coluna(df, [col])
        if col_real:
            df[col_real] = pd.to_datetime(df[col_real], dayfirst=True, errors="coerce")
            log.append(f"[OK] Datetime: '{col_real}'")

    for col in cfg.colunas_valor:
        col_real = _detectar_coluna(df, [col])
        if col_real:
            df[col_real] = df[col_real].apply(_limpar_numero).astype("float64")
            log.append(f"[OK] Float: '{col_real}'")

    df.reset_index(drop=True, inplace=True)
    log.append(f"[CONCLUÍDO] Base final: {df.shape[0]} linhas × {df.shape[1]} colunas")

    return df, log


def exportar_excel(df: pd.DataFrame, caminho_saida: str = "Base_Vendas_Limpa.xlsx") -> str:
    """Exporta o DataFrame para Excel com formatação profissional."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    with pd.ExcelWriter(caminho_saida, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Base_Limpa")
        ws = writer.sheets["Base_Limpa"]

        fill   = PatternFill("solid", start_color="1E3A5F")
        fonte  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
        borda  = Border(bottom=Side(style="medium", color="2563EB"))
        centro = Alignment(horizontal="center", vertical="center")

        for cell in ws[1]:
            cell.fill = fill
            cell.font = fonte
            cell.alignment = centro
            cell.border = borda

        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)

        ws.row_dimensions[1].height = 22

    return caminho_saida


# ─── WATCHDOG ─────────────────────────────────────────────────────────────

def iniciar_watchdog(
    pasta_entrada: str = "./input",
    pasta_saida: str   = "./output",
    config: Optional[ETLConfig] = None,
) -> None:
    """
    Monitora uma pasta e processa automaticamente qualquer .xlsx/.xls/.csv
    assim que um novo arquivo é detectado.

    Requer: pip install watchdog
    """
    try:
        from watchdog.observers import Observer
        from watchdog.events import FileSystemEventHandler
    except ImportError:
        print("[ERRO] Instale o watchdog: pip install watchdog")
        sys.exit(1)

    import time

    Path(pasta_entrada).mkdir(parents=True, exist_ok=True)
    Path(pasta_saida).mkdir(parents=True, exist_ok=True)

    class ETLHandler(FileSystemEventHandler):
        def on_created(self, event):
            if event.is_directory:
                return
            path = Path(event.src_path)
            if path.suffix.lower() not in (".xlsx", ".xls", ".csv"):
                return
            # Aguarda o arquivo fechar (escrita completa)
            time.sleep(1)
            print(f"\n[WATCHDOG] Novo arquivo detectado: {path.name}")
            with open(path, "rb") as f:
                bytes_arq = f.read()
            df, log = rodar_pipeline(bytes_arq, path.name, config)
            for linha in log:
                print(f"  {linha}")
            if df is not None:
                saida = Path(pasta_saida) / f"LIMPO_{path.stem}.xlsx"
                exportar_excel(df, str(saida))
                print(f"[WATCHDOG] Exportado: {saida}")

    observer = Observer()
    observer.schedule(ETLHandler(), pasta_entrada, recursive=False)
    observer.start()
    print(f"[WATCHDOG] Monitorando '{pasta_entrada}' → '{pasta_saida}'")
    print("[WATCHDOG] Pressione Ctrl+C para parar.\n")
    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        observer.stop()
    observer.join()


# ─── LAMBDA / CLOUD FUNCTION HANDLER ────────────────────────────────────

def lambda_handler(event: dict, context=None) -> dict:
    """
    Handler para AWS Lambda ou Google Cloud Function.

    Espera um evento com:
    {
        "file_base64": "<base64 do arquivo>",
        "file_name": "vendas.xlsx",
        "config": { ... }   ← opcional
    }

    Retorna:
    {
        "statusCode": 200,
        "body": { "rows": N, "columns": [...], "log": [...] }
    }

    Para uso com S3: leia o arquivo do bucket e passe os bytes.
    Para uso com GCS: idem com google.cloud.storage.
    """
    import base64

    try:
        file_b64  = event.get("file_base64", "")
        file_name = event.get("file_name", "upload.xlsx")
        cfg_dict  = event.get("config", {})

        arquivo_bytes = base64.b64decode(file_b64)
        config = ETLConfig(**cfg_dict) if cfg_dict else ETLConfig()

        df, log = rodar_pipeline(arquivo_bytes, file_name, config)

        if df is None:
            return {"statusCode": 422, "body": {"error": "ETL falhou", "log": log}}

        # Serializa a saída como JSON (ou base64 do Excel)
        buf = io.BytesIO()
        df.to_csv(buf, index=False)
        csv_b64 = base64.b64encode(buf.getvalue()).decode()

        return {
            "statusCode": 200,
            "body": {
                "rows": len(df),
                "columns": list(df.columns),
                "log": log,
                "csv_base64": csv_b64,
            },
        }
    except Exception as e:
        return {"statusCode": 500, "body": {"error": str(e)}}


# ─── CLI ──────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python etl_engine.py <arquivo.xlsx> [saida.xlsx] [config.json]")
        print("     python etl_engine.py --watch <pasta_entrada> <pasta_saida>")
        sys.exit(0)

    if sys.argv[1] == "--watch":
        pasta_in  = sys.argv[2] if len(sys.argv) > 2 else "./input"
        pasta_out = sys.argv[3] if len(sys.argv) > 3 else "./output"
        iniciar_watchdog(pasta_in, pasta_out)
    else:
        caminho_entrada = sys.argv[1]
        caminho_saida   = sys.argv[2] if len(sys.argv) > 2 else "Base_Vendas_Limpa.xlsx"
        config_path     = sys.argv[3] if len(sys.argv) > 3 else None

        config = ETLConfig.from_json(config_path) if config_path else ETLConfig()

        with open(caminho_entrada, "rb") as f:
            bytes_arq = f.read()

        df, log = rodar_pipeline(bytes_arq, Path(caminho_entrada).name, config)

        for linha in log:
            print(linha)

        if df is not None:
            saida = exportar_excel(df, caminho_saida)
            print(f"\n✅ Arquivo exportado: {saida}")